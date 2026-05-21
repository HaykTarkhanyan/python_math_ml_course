"""Generate the focused Metric ML & GenAI study-tracker Excel workbook.

13 essential topics (4 methodology + 4 GenAI core + 5 GenAI practical).
Classical models and the NN/backprop/gradient-descent bridge are assumed
background for the intended audience, so they are intentionally left out.

Edit the R / TRACKER / START_HERE data below and re-run to regenerate.
Output: misc/Metric_Prep_ML_GenAI_Tracker.xlsx
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------- logging
LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_DIR / "build_prep_excel.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

OUT = Path("misc") / "Metric_Prep_ML_GenAI_Tracker.xlsx"

# ---------------------------------------------------------------- palette
RED = "D90012"      # Armenian flag red    -> headers + GenAI-practical accent
BLUE = "0033A0"     # Armenian flag blue   -> Methodology accent
ORANGE = "F2A800"   # Armenian flag orange -> GenAI-core accent
LINK = "0563C1"

FILL_HEADER = PatternFill("solid", fgColor=RED)
FILL_TITLE = PatternFill("solid", fgColor=RED)
FILL_NOT = PatternFill("solid", fgColor="D9D9D9")   # grey
FILL_LEARN = PatternFill("solid", fgColor="FBE2B7")  # light orange
FILL_CONF = PatternFill("solid", fgColor="CBD7F0")   # light blue
FILL_MAST = PatternFill("solid", fgColor="C6EFCE")   # light green

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(vertical="top", wrap_text=True)
WRAP_CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)

STATUSES = ["Not started", "Learning", "Confident", "Mastered"]

CLUSTER_COLOR = {
    "Methodology": BLUE,
    "GenAI core": ORANGE,
    "GenAI practical": RED,
}

# ---------------------------------------------------------------- resources
# key -> (label, url). Verified from the .md file, the user's message,
# Hugging Face docs, and a few well-known landing pages.
R = {
    "ng": ("Andrew Ng ML Specialization", "https://www.coursera.org/specializations/machine-learning-introduction"),
    "sq_ml": ("StatQuest: ML playlist", "https://www.youtube.com/playlist?list=PLblh5JKOoLUIcdlgu78MnlATeyx4cEVeR"),
    "sq_ch": ("StatQuest channel", "https://www.youtube.com/c/joshstarmer"),
    "d2l": ("Dive into Deep Learning", "https://d2l.ai/"),
    "hf": ("Hugging Face Learn", "https://huggingface.co/learn/"),
    "3b1b_nn": ("3Blue1Brown: Neural Networks", "https://www.youtube.com/playlist?list=PLZHQObOWTQDNU6R1_67000Dx_ZCJB-3pi"),
    "3b1b_attn": ("3Blue1Brown: Attention", "https://www.youtube.com/watch?v=eMlx5fFNoYc"),
    "karp_gpt": ("Karpathy: Let's build GPT", "https://www.youtube.com/watch?v=kCc8FmEb1nY"),
    "karp_llm": ("Karpathy: Deep Dive into LLMs", "https://www.youtube.com/watch?v=7xTGNNLPyMI"),
    "karp_tok": ("Karpathy: Build the GPT Tokenizer", "https://www.youtube.com/watch?v=zduSFxRajkE"),
    "ill_tf": ("Illustrated Transformer (Alammar)", "https://jalammar.github.io/illustrated-transformer/"),
    "ill_w2v": ("Illustrated Word2vec (Alammar)", "https://jalammar.github.io/illustrated-word2vec/"),
    "sq_w2v": ("StatQuest: Word2Vec", "https://youtu.be/viZrOnJclY0"),
    "raschka_attn": ("Coding Self-Attention (Raschka)", "https://magazine.sebastianraschka.com/p/understanding-and-coding-self-attention"),
    "kazem_pe": ("Positional Encoding (Kazemnejad)", "https://kazemnejad.com/blog/transformer_architecture_positional_encoding/"),
    "hf_bpe": ("BPE tokenization (HF Ch6)", "https://huggingface.co/learn/llm-course/en/chapter6/5"),
    "raschka_ft": ("Finetuning LLMs (Raschka)", "https://magazine.sebastianraschka.com/p/finetuning-large-language-models"),
    "hf_trainer": ("Fine-tuning w/ Trainer API (HF Ch3)", "https://huggingface.co/learn/llm-course/en/chapter3/3"),
    "hf_generate": ("How to generate text (HF)", "https://huggingface.co/blog/how-to-generate"),
    "labonne_dec": ("Decoding Strategies (Labonne)", "https://huggingface.co/blog/mlabonne/decoding-strategies"),
    "promptguide": ("Prompt Engineering Guide (DAIR.AI)", "https://www.promptingguide.ai/"),
    "skl_cv": ("scikit-learn: Cross-validation", "https://scikit-learn.org/stable/modules/cross_validation.html"),
    "skl_grid": ("scikit-learn: Grid Search", "https://scikit-learn.org/stable/modules/grid_search.html"),
    "skl_metrics": ("scikit-learn: Metrics", "https://scikit-learn.org/stable/modules/model_evaluation.html"),
    "optuna": ("Optuna (hyperparameter optimization)", "https://optuna.org/"),
    "rag_paper": ("RAG paper (Lewis et al., 2020)", "https://arxiv.org/abs/2005.11401"),
    "hf_optim": ("Optimizing LLMs for Speed & Memory (HF)", "https://huggingface.co/docs/transformers/llm_tutorial_optimization"),
    "vllm": ("vLLM docs (serving / KV-cache)", "https://docs.vllm.ai/"),
    "course": ("Metric course (Python/Math/ML)", "https://hayktarkhanyan.github.io/python_math_ml_course/"),
    "welch": ("Welch Labs channel", "https://www.youtube.com/@WelchLabsVideo"),
}

# (cluster, topic, [resource keys, up to 2], notes)
TRACKER = [
    # ---- Methodology -----------------------------------------------------
    ("Methodology", "Performance metrics (Accuracy, Precision, Recall, F1, ROC, AUC)",
     ["sq_ml", "skl_metrics"],
     "Confusion matrix, precision/recall trade-off, F1, ROC & AUC. StatQuest has a crisp video for each."),
    ("Methodology", "Bias-variance tradeoff",
     ["sq_ml", "ng"],
     "Overfitting vs underfitting; regularization (L1/L2, dropout) is the lever."),
    ("Methodology", "Cross-validation",
     ["sq_ml", "skl_cv"],
     "k-fold, stratified, leave-one-out; why a single train/test split can mislead."),
    ("Methodology", "Hyperparameter tuning",
     ["skl_grid", "optuna"],
     "Grid/Randomized search; Optuna for smarter Bayesian/TPE search."),

    # ---- GenAI core ------------------------------------------------------
    ("GenAI core", "Word embeddings (Word2Vec: CBOW, Skip-gram)",
     ["ill_w2v", "sq_w2v"],
     "Dense vectors; CBOW vs Skip-gram; word analogies (king - man + woman)."),
    ("GenAI core", "Attention mechanism",
     ["3b1b_attn", "raschka_attn"],
     "The core idea before full transformers. Illustrated Transformer (Alammar) shows it in context."),
    ("GenAI core", "Transformer architecture (Self-/Multi-Head Attention)",
     ["ill_tf", "karp_gpt"],
     "Positional encoding -> Kazemnejad blog; 3B1B 'Attention' video dissects the internals; Karpathy builds one in code."),
    ("GenAI core", "Tokenization (BPE, WordPiece, SentencePiece)",
     ["hf_bpe", "karp_tok"],
     "BPE (GPT), WordPiece (BERT), SentencePiece (T5). Karpathy builds a tokenizer from scratch."),

    # ---- GenAI practical -------------------------------------------------
    ("GenAI practical", "Prompting (Zero-shot, Few-shot, Chain-of-Thought)",
     ["promptguide", "karp_llm"],
     "Highest-leverage skill: getting a model to perform without training it."),
    ("GenAI practical", "Fine-tuning (SFT, LoRA, Adapters)",
     ["raschka_ft", "hf_trainer"],
     "Full fine-tuning vs PEFT. LoRA/adapters via HF PEFT library. See the 'fine-tune a small LM' project."),
    ("GenAI practical", "Decoding strategies (Greedy, Beam, Top-k, Top-p, Temperature)",
     ["hf_generate", "labonne_dec"],
     "How tokens are actually sampled at generation time; temperature controls randomness."),
    ("GenAI practical", "RAG (Retrieval-Augmented Generation)",
     ["hf", "rag_paper"],
     "Retriever + vector DB + generator; grounds answers in your documents. Hands-on: build a simple RAG pipeline."),
    ("GenAI practical", "Inference optimization (Quantization, KV-Cache, Distillation)",
     ["hf_optim", "vllm"],
     "Quantization (bitsandbytes/AWQ, 4/8-bit), KV-cache, Flash Attention, distillation. vLLM = the standard serving engine."),
]

# (order, label, resource_key, instructor note)
START_HERE = [
    (1, "3Blue1Brown - Neural Networks playlist", "3b1b_nn",
     "Watch all of these if you haven't yet - especially the transformer ones. I've watched the series ~3 times at different stages and something new clicks every time."),
    (2, "Andrej Karpathy - Let's build GPT: from scratch, in code, spelled out", "karp_gpt",
     "Watch this next - it immediately makes the practical side click."),
    (3, "Andrej Karpathy - Deep Dive into LLMs like ChatGPT", "karp_llm",
     "A bit long, but anything Karpathy makes is fantastic - almost popcorn-watching."),
]

# broad references for the Legend sheet
BROAD = ["course", "d2l", "hf", "sq_ch", "3b1b_nn", "welch"]
USE_FOR = {
    "course": "This very course (Python, Math, ML) in Armenian",
    "d2l": "Deep learning with runnable code - the workhorse reference",
    "hf": "NLP, LLMs, Transformers - hands-on courses",
    "sq_ch": "Crystal-clear explainers for almost every ML topic",
    "3b1b_nn": "Visual intuition for neural nets + transformers",
    "welch": "Deep, beautiful intuition videos",
}


def link_cell(ws, row, col, key):
    """Write resource `key` into a cell as a clickable hyperlink."""
    if key is None:
        ws.cell(row=row, column=col).border = BORDER
        return
    label, url = R[key]
    c = ws.cell(row=row, column=col, value=label)
    c.hyperlink = url
    c.font = Font(color=LINK, underline="single", size=10)
    c.alignment = WRAP
    c.border = BORDER


def title_block(ws, text, subtitle, span):
    ws["A1"] = text
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells(f"A1:{span}1")
    ws.row_dimensions[1].height = 26
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="555555")
    ws["A2"].alignment = WRAP
    ws.merge_cells(f"A2:{span}2")
    ws.row_dimensions[2].height = 32


def build_start_here(wb):
    ws = wb.active
    ws.title = "Start Here"
    ws.sheet_view.showGridLines = False
    title_block(ws, "Metric Prep - Start Here",
                "Best entry points, in order. The notes are the instructor's personal recommendation. "
                "Then work through the 'Study Tracker' tab topic by topic.", "D")

    for j, h in enumerate(["#", "Watch", "Link", "Why (instructor note)"], start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = FILL_HEADER
        c.alignment = WRAP_CENTER
        c.border = BORDER

    r = 5
    for order, label, key, note in START_HERE:
        ws.cell(row=r, column=1, value=order).alignment = WRAP_CENTER
        ws.cell(row=r, column=1).border = BORDER
        tc = ws.cell(row=r, column=2, value=label)
        tc.alignment = WRAP
        tc.font = Font(bold=True)
        tc.border = BORDER
        link_cell(ws, r, 3, key)
        nc = ws.cell(row=r, column=4, value=note)
        nc.alignment = WRAP
        nc.border = BORDER
        ws.row_dimensions[r].height = 60
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Tip").font = Font(bold=True, color=RED)
    tip = ws.cell(row=r, column=2,
                  value="Anything by 3Blue1Brown, Andrej Karpathy, and Welch Labs is fantastic for building deep intuition.")
    tip.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

    for col, w in zip("ABCD", [5, 46, 34, 60]):
        ws.column_dimensions[col].width = w


def build_tracker(wb):
    ws = wb.create_sheet("Study Tracker")
    ws.sheet_view.showGridLines = False
    title_block(ws, "ML & GenAI - Study Tracker (focused)",
                "Set each topic's Status from the dropdown - cells auto-color and the counters update. "
                "13 essential topics; classical ML + NN/backprop are assumed background.", "F")

    n = len(TRACKER)
    header_row = 7
    first = header_row + 1
    last = header_row + n
    status_range = f"C{first}:C{last}"

    # KPI labels (row 4) + values (row 5)
    labels = ["Not started", "Learning", "Confident", "Mastered", "% Confident+"]
    for j, lab in enumerate(labels, start=1):
        c = ws.cell(row=4, column=j, value=lab)
        c.font = Font(bold=True, color="555555")
        c.alignment = WRAP_CENTER
    for j, name in enumerate(STATUSES, start=1):
        v = ws.cell(row=5, column=j, value=f'=COUNTIF({status_range},"{name}")')
        v.font = Font(bold=True, size=12, color=BLUE)
        v.alignment = WRAP_CENTER
    pct = ws.cell(row=5, column=5, value=f"=(C5+D5)/{n}")
    pct.number_format = "0%"
    pct.font = Font(bold=True, size=12, color=RED)
    pct.alignment = WRAP_CENTER

    headers = ["Cluster", "Topic", "Status", "Resource 1", "Resource 2", "Notes"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = FILL_HEADER
        c.alignment = WRAP_CENTER
        c.border = BORDER

    for i, (cluster, topic, keys, notes) in enumerate(TRACKER):
        row = first + i
        cc = ws.cell(row=row, column=1, value=cluster)
        cc.font = Font(bold=True, color=CLUSTER_COLOR[cluster])
        cc.alignment = WRAP_CENTER
        cc.border = BORDER

        tc = ws.cell(row=row, column=2, value=topic)
        tc.alignment = WRAP
        tc.border = BORDER

        sc = ws.cell(row=row, column=3, value="Not started")
        sc.alignment = WRAP_CENTER
        sc.border = BORDER

        link_cell(ws, row, 4, keys[0] if len(keys) > 0 else None)
        link_cell(ws, row, 5, keys[1] if len(keys) > 1 else None)

        nc = ws.cell(row=row, column=6, value=notes)
        nc.alignment = WRAP
        nc.font = Font(size=10, color="555555")
        nc.border = BORDER
        ws.row_dimensions[row].height = 44

    dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUSES),
                        allow_blank=True, showErrorMessage=True)
    dv.error = "Pick one of: " + ", ".join(STATUSES)
    dv.prompt = "Choose your level for this topic"
    ws.add_data_validation(dv)
    dv.add(status_range)

    for value, fill in [("Not started", FILL_NOT), ("Learning", FILL_LEARN),
                        ("Confident", FILL_CONF), ("Mastered", FILL_MAST)]:
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill),
        )

    ws.auto_filter.ref = f"A{header_row}:F{last}"
    ws.freeze_panes = f"A{first}"

    for col, w in zip("ABCDEF", [16, 44, 13, 30, 30, 50]):
        ws.column_dimensions[col].width = w


def build_legend(wb):
    ws = wb.create_sheet("Legend & Broad Resources")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Legend & Broad Resources"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 26

    ws["A3"] = "Status colors"
    ws["A3"].font = Font(bold=True, size=12, color=RED)
    r = 4
    for name, fill in [("Not started", FILL_NOT), ("Learning", FILL_LEARN),
                       ("Confident", FILL_CONF), ("Mastered", FILL_MAST)]:
        c = ws.cell(row=r, column=1, value=name)
        c.fill = fill
        c.border = BORDER
        c.alignment = WRAP_CENTER
        r += 1

    ws.cell(row=r + 1, column=1, value="Cluster colors: Methodology (blue), GenAI core (orange), GenAI practical (red)").font = Font(italic=True, color="777777")

    start = r + 3
    ws.cell(row=start, column=1, value="Broad resources (good for any topic)").font = Font(bold=True, size=12, color=RED)
    hdr = start + 1
    for j, h in enumerate(["Resource", "Link", "Use for"], start=1):
        c = ws.cell(row=hdr, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = FILL_HEADER
        c.border = BORDER
        c.alignment = WRAP_CENTER

    rr = hdr + 1
    for key in BROAD:
        nc = ws.cell(row=rr, column=1, value=R[key][0])
        nc.font = Font(bold=True)
        nc.alignment = WRAP
        nc.border = BORDER
        link_cell(ws, rr, 2, key)
        uc = ws.cell(row=rr, column=3, value=USE_FOR.get(key, ""))
        uc.alignment = WRAP
        uc.border = BORDER
        ws.row_dimensions[rr].height = 30
        rr += 1

    ws.cell(row=rr + 2, column=1,
            value="Built from misc/Metric Preparation Material.md + hand-picked links.").font = Font(italic=True, color="777777")

    for col, w in zip("ABC", [34, 36, 52]):
        ws.column_dimensions[col].width = w


def main():
    wb = Workbook()
    build_start_here(wb)
    build_tracker(wb)
    build_legend(wb)
    wb.save(OUT)
    log.info("Wrote %s (%d topics)", OUT, len(TRACKER))


if __name__ == "__main__":
    main()
