"""Generate 01_calc.xlsx as a student-facing study sheet.

Run: uv run --with openpyxl python misc/solo/make_calc_xlsx.py
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


Path("logs").mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("logs/make_calc_xlsx.log"),
    ],
)
log = logging.getLogger(__name__)


OUT = Path("misc/solo/01_calc.xlsx")

HEADERS = ["Section", "Item", "Link", "Teacher's note", "Done", "Student's question / comment"]

# Light tints of the Armenian flag colors for section banding
COLOR_RED = "F9D5D8"
COLOR_BLUE = "D6DBEC"
COLOR_ORANGE = "FCEBC8"

SEC_EXT = "Extrema, convexity, Taylor series"
SEC_INT = "Integrals"
SEC_MUL = "Multivariate calculus (gradients, Jacobians, Hessians)"

SECTION_COLOR = {
    SEC_EXT: COLOR_RED,
    SEC_INT: COLOR_BLUE,
    SEC_MUL: COLOR_ORANGE,
}

# (section, item, link, teacher_note)
ROWS = [
    (SEC_EXT, "Chapter", "https://hayktarkhanyan.github.io/python_math_ml_course/math/05_calc_extrema_convexity_taylor.html",
     "This will be quite important for the optimization section of the course."),
    (SEC_EXT, "Lecture video", "https://youtu.be/OHXH7zn65zU", ""),
    (SEC_EXT, "Practical video", "https://youtu.be/d5bAQnILhvs", ""),
    (SEC_EXT, "HW 02", "", ""),
    (SEC_EXT, "HW 03", "", ""),
    (SEC_EXT, "HW 04", "", ""),
    (SEC_EXT, "HW 05", "",
     "There is some extra terminology (e.g. regularizer parameter) - it's safe to ignore it for now, just focus on the math component."),
    (SEC_EXT, "HW 08", "", ""),

    (SEC_INT, "Chapter", "https://hayktarkhanyan.github.io/python_math_ml_course/math/06_calc_integrals.html",
     "Exact formulas for calculations (e.g. integration by parts) are not important for ML, but the general idea of what an integral represents is quite important."),
    (SEC_INT, "Lecture video", "https://youtu.be/J2tU1cUlZok", ""),
    (SEC_INT, "Practical video", "https://youtu.be/FRjc1KiAbaw", ""),
    (SEC_INT, "HW 01", "", "Safe to skip some boring calculations."),
    (SEC_INT, "HW 02", "", ""),
    (SEC_INT, "HW 04", "", "The convolution here is foreshadowing the convolution operation which we'll cover in far future lectures."),
    (SEC_INT, "HW 05", "", "This is more of a brain teaser, and a cool thing to derive yourself, but not super important for ML."),

    (SEC_MUL, "Chapter", "https://hayktarkhanyan.github.io/python_math_ml_course/math/07_calc_multivar.html", ""),
    (SEC_MUL, "Video: multivar func, gradient descent", "https://youtu.be/tFiDNbpZNFw", ""),
    (SEC_MUL, "Video: extrema, hessian", "https://youtu.be/Fc8lIjjDE5U", ""),
    (SEC_MUL, "Practical video", "https://youtu.be/XFF03oPjAyE", ""),
    (SEC_MUL, "HW 01", "", "Good opportunity to give the house price project we talked about today a new try."),
    (SEC_MUL, "HW 02", "", ""),
    (SEC_MUL, "HW 03", "", ""),
    (SEC_MUL, "HW 04, 05, 06", "",
     "These are more computation heavy. If you feel comfortable with the topic, no need to fully do them."),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculus"

    # header row
    header_fill = PatternFill("solid", fgColor="333333")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # data rows
    for i, (section, item, link, note) in enumerate(ROWS, start=2):
        fill = PatternFill("solid", fgColor=SECTION_COLOR[section])

        ws.cell(row=i, column=1, value=section)
        ws.cell(row=i, column=2, value=item)

        link_cell = ws.cell(row=i, column=3, value=link)
        if link:
            link_cell.hyperlink = link
            link_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=i, column=4, value=note)
        ws.cell(row=i, column=5, value="")  # Done
        ws.cell(row=i, column=6, value="")  # Student's question/comment

        for col in range(1, 7):
            c = ws.cell(row=i, column=col)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # column widths
    widths = {1: 38, 2: 32, 3: 55, 4: 50, 5: 8, 6: 45}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # checkbox-style dropdown on Done column
    dv = DataValidation(type="list", formula1='"✓"', allow_blank=True)
    dv.add(f"E2:E{len(ROWS) + 1}")
    ws.add_data_validation(dv)

    # freeze header
    ws.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    log.info(f"Wrote {OUT.resolve()} ({len(ROWS)} rows)")


if __name__ == "__main__":
    main()
