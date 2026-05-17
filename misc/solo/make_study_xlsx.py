"""Generate study_plan.xlsx with multiple topic sheets for student supervision.

Run: uv run --with openpyxl python misc/solo/make_study_xlsx.py
"""

import csv
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


Path("logs").mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("logs/make_study_xlsx.log"),
    ],
)
log = logging.getLogger(__name__)


OUT = Path("misc/solo/study_plan.xlsx")
OPTIM_CSV = Path("misc/solo/optimization.csv")

HEADERS = ["Section", "Item", "Link", "Note", "Done", "Questions / comments"]

# Light tints of the Armenian flag palette
COLOR_RED = "F9D5D8"
COLOR_BLUE = "D6DBEC"
COLOR_ORANGE = "FCEBC8"


# --- Calculus sheet ---
SEC_EXT = "Extrema, convexity, Taylor series"
SEC_INT = "Integrals"
SEC_MUL = "Multivariate calculus (gradients, Jacobians, Hessians)"

CALC_COLORS = {
    SEC_EXT: COLOR_RED,
    SEC_INT: COLOR_BLUE,
    SEC_MUL: COLOR_ORANGE,
}

CALC_ROWS = [
    (SEC_EXT, "Chapter", "https://hayktarkhanyan.github.io/python_math_ml_course/math/05_calc_extrema_convexity_taylor.html",
     "This will be quite important for the optimization section of the course."),
    (SEC_EXT, "Lecture video", "https://youtu.be/OHXH7zn65zU", ""),
    (SEC_EXT, "Practical video", "https://youtu.be/d5bAQnILhvs", ""),
    (SEC_EXT, "HW 02", "", ""),
    (SEC_EXT, "HW 03", "", ""),
    (SEC_EXT, "HW 04", "", ""),
    (SEC_EXT, "HW 05", "",
     "There is some extra terminology (e.g. regularizer parameter) - it's safe to ignore it for now, just focus on the math component."),
    (SEC_EXT, "HW 08", "", ""),

    (SEC_INT, "Chapter", "https://hayktarkhanyan.github.io/python_math_ml_course/math/06_calc_integrals.html",
     "Exact formulas for calculations (e.g. integration by parts) are not important for ML, but the general idea of what an integral represents is quite important."),
    (SEC_INT, "Lecture video", "https://youtu.be/J2tU1cUlZok", ""),
    (SEC_INT, "Practical video", "https://youtu.be/FRjc1KiAbaw", ""),
    (SEC_INT, "HW 01", "", "Safe to skip some boring calculations."),
    (SEC_INT, "HW 02", "", ""),
    (SEC_INT, "HW 04", "", "The convolution here is foreshadowing the convolution operation which we'll cover in far future lectures."),
    (SEC_INT, "HW 05", "", "This is more of a brain teaser, and a cool thing to derive yourself, but not super important for ML."),

    (SEC_MUL, "Chapter", "https://hayktarkhanyan.github.io/python_math_ml_course/math/07_calc_multivar.html", ""),
    (SEC_MUL, "Video: multivar func, gradient descent", "https://youtu.be/tFiDNbpZNFw", ""),
    (SEC_MUL, "Video: extrema, hessian", "https://youtu.be/Fc8lIjjDE5U", ""),
    (SEC_MUL, "Practical video", "https://youtu.be/XFF03oPjAyE", ""),
    (SEC_MUL, "HW 01", "", "Good opportunity to give the house price project we talked about today a new try."),
    (SEC_MUL, "HW 02", "", ""),
    (SEC_MUL, "HW 03", "", ""),
    (SEC_MUL, "HW 04, 05, 06", "",
     "These are more computation heavy. If you feel comfortable with the topic, no need to fully do them."),
]


# --- Optimization sheet ---
SEC_GD = "Gradient descent: prerequisites & vanilla GD"
SEC_MOMENTUM = "Momentum, Adam, and friends"

OPTIM_COLORS = {
    SEC_GD: COLOR_BLUE,
    SEC_MOMENTUM: COLOR_ORANGE,
}

# HW content lives in the chapter qmd; rows below just track progress per problem.
OPTIM_ROWS = [
    (SEC_GD, "Chapter", "https://hayktarkhanyan.github.io/python_math_ml_course/math/09_optim_prereq__gradient_descent.html",
     "Builds on multivariate calculus (gradients, Hessians) and Taylor series. Big new idea: iterative optimization - no more closed-form solutions."),
    (SEC_GD, "Lecture 15 — Prerequisites, 1st/2nd order conditions", "https://youtu.be/7bl1k6TZ9_Q", ""),
    (SEC_GD, "Lecture 16 — (In)Exact Line Search, GD intro", "https://youtu.be/UjGD31i8Ji4", ""),
    (SEC_GD, "Practical — GD learning rate schedule", "https://youtu.be/H8M6p-uSxCM", ""),
    (SEC_GD, "HW 01 — Condition number by hand", "", "Pen-and-paper warmup."),
    (SEC_GD, "HW 02 — Implement vanilla GD from scratch", "", "numpy only, no sklearn or torch. Build the loop yourself."),
    (SEC_GD, "HW 03 — Find the divergence threshold", "", "Experimental — just trial and error with the learning rate."),
    (SEC_GD, "HW 04 — Beat constant LR with a custom schedule", "", "There's a starter cell already in Lectures/optim/03_gd_step_size.ipynb."),
    (SEC_GD, "HW 05 — Visualize the zig-zag", "", "Picture is way more convincing than the math alone."),
    (SEC_GD, "HW 06 — Stuck at a saddle?", "", "Tiny perturbation in starting point changes everything - the whole point of the problem."),
    (SEC_GD, "HW 07 — GD for linear regression", "", "First time applying GD to real data. Compare against the closed-form solution."),

    (SEC_MOMENTUM, "Chapter", "https://hayktarkhanyan.github.io/python_math_ml_course/math/10_optim_momentum_first_order_algs.html",
     "Builds on chapter 09. Adam/RMSProp/Adagrad fix vanilla GD's two big weaknesses: zig-zagging on ill-conditioned problems, and slow convergence."),
    (SEC_MOMENTUM, "Lecture 17 — Momentum, ADAM, RMSProp, Adagrad", "https://youtu.be/JsSnSHZtG_o", ""),
    (SEC_MOMENTUM, "Practical 11 — Adam and friends in optimization", "https://youtu.be/mGOCt9ZcWJg", ""),
    (SEC_MOMENTUM, "HW 01 — Write out the update rules", "", "Recall from lecture / notes, then verify against the notebook."),
    (SEC_MOMENTUM, "HW 02 — Implement momentum from scratch", "", "Builds on chapter 09 HW 02 (vanilla GD)."),
    (SEC_MOMENTUM, "HW 03 — Momentum kills the zig-zag", "", "Side-by-side with chapter 09 HW 05. Visual punchline."),
    (SEC_MOMENTUM, "HW 04 — Implement Adam from scratch", "", "Most substantial implementation in the chapter. Follow the formulas exactly."),
    (SEC_MOMENTUM, "HW 05 — The bias correction matters", "", "Quick experiment - the 'why' explanation is the key part."),
    (SEC_MOMENTUM, "HW 06 — When does Adam fail?", "", "Counter-example to 'always use Adam'."),
    (SEC_MOMENTUM, "HW 07 — Robustness to the learning rate", "", "Adam's main practical sell. 5 LRs spanning 4 orders of magnitude."),
    (SEC_MOMENTUM, "HW 08 — Adam for linear regression", "", "Builds on chapter 09 HW 07. Often surprising: Adam doesn't help on convex problems."),
]


def build_sheet(wb: Workbook, title: str, rows: list, section_colors: dict) -> None:
    """Build one sheet with the standard layout."""
    ws = wb.create_sheet(title)

    # header
    header_fill = PatternFill("solid", fgColor="333333")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # data rows
    for i, (section, item, link, note) in enumerate(rows, start=2):
        fill = PatternFill("solid", fgColor=section_colors[section])

        ws.cell(row=i, column=1, value=section)
        ws.cell(row=i, column=2, value=item)

        link_cell = ws.cell(row=i, column=3, value=link)
        if link:
            link_cell.hyperlink = link
            link_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=i, column=4, value=note)
        ws.cell(row=i, column=5, value="")
        ws.cell(row=i, column=6, value="")

        for col in range(1, 7):
            c = ws.cell(row=i, column=col)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # column widths
    widths = {1: 38, 2: 40, 3: 55, 4: 60, 5: 8, 6: 45}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # checkbox-style dropdown on Done column
    dv = DataValidation(type="list", formula1='"✓"', allow_blank=True)
    dv.add(f"E2:E{len(rows) + 1}")
    ws.add_data_validation(dv)

    # freeze header
    ws.freeze_panes = "A2"


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop default empty sheet

    build_sheet(wb, "Calculus", CALC_ROWS, CALC_COLORS)
    build_sheet(wb, "Optimization", OPTIM_ROWS, OPTIM_COLORS)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    log.info(f"Wrote {OUT.resolve()} (calc={len(CALC_ROWS)} rows, optim={len(OPTIM_ROWS)} rows)")

    # standalone CSV of the optimization sheet (utf-8-sig for Excel compat)
    with OPTIM_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for section, item, link, note in OPTIM_ROWS:
            w.writerow([section, item, link, note, "", ""])
    log.info(f"Wrote {OPTIM_CSV.resolve()} ({len(OPTIM_ROWS)} rows)")


if __name__ == "__main__":
    main()
