"""Add a "Probability, Stats & Extras" overview sheet to study_plan_2026.xlsx.

Loads the existing workbook (preserving the student's filled-in Calculus and
Optimization sheets), drops the scratch "Sheet2", and adds a roadmap sheet for
chapters 08, 11-15, 16-20, 21-26, 28, 29, 31 with per-chapter priority notes.

Idempotent: re-running replaces the overview sheet without touching the others.

Run: uv run --with openpyxl python misc/solo/add_prob_stats_sheet.py
"""

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


Path("logs").mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("logs/add_prob_stats_sheet.log"),
    ],
)
log = logging.getLogger(__name__)


XLSX = Path("misc/solo/study_plan_2026.xlsx")
SHEET_NAME = "Probability, Stats and Extras"  # no ampersand: Google Sheets escapes & to &amp;

HEADERS = ["Section", "Item", "Link", "Note", "Done", "Question / comment"]

# Light tints (consistent with the existing sheets' banding)
COLOR_ORANGE = "FCEBC8"
COLOR_BLUE = "D6DBEC"
COLOR_RED = "F9D5D8"
COLOR_GREEN = "D9EAD3"

BASE = "https://hayktarkhanyan.github.io/python_math_ml_course/math/"

SEC_OPT = "Optimization extras (optional)"
SEC_PROB = "Probability (important - core for ML)"
SEC_STAT = "Statistics"
SEC_INFO = "Information theory & curse of dimensionality"

SECTION_COLOR = {
    SEC_OPT: COLOR_ORANGE,
    SEC_PROB: COLOR_BLUE,
    SEC_STAT: COLOR_RED,
    SEC_INFO: COLOR_GREEN,
}

# (section, item, link, note)
ROWS = [
    (SEC_OPT, "08 Univariate optimization", BASE + "08_optim_univar.html",
     "Not super important, but can be interesting since you came up with the idea of this optimization approach yourself."),
    (SEC_OPT, "11 Second-order methods", BASE + "11_optim_second_order.html",
     "All of 11-15 are optional. Fine to skip for now - do them only if you enjoy optimization."),
    (SEC_OPT, "12 Derivative-free optimization", BASE + "12_derivative_free.html", "Optional."),
    (SEC_OPT, "13 Evolutionary algorithms", BASE + "13_evolutionary.html", "Optional."),
    (SEC_OPT, "14 Bayesian optimization", BASE + "14_bayesian.html",
     "Optional now. We may revisit this in the future, but it's very low priority right now."),
    (SEC_OPT, "15 Multicriteria optimization", BASE + "15_multicriteria_optimization.html", "Optional."),

    (SEC_PROB, "16 Probability intro", BASE + "16_probability_intro.html",
     "The whole probability block matters. For homeworks across these chapters: feel free to skip ones that feel trivial or boring - the ideas are what's important."),
    (SEC_PROB, "17 Expectation, variance, inequalities", BASE + "17_probability_exp_var_inequalities.html", ""),
    (SEC_PROB, "18 Correlation & covariance", BASE + "18_probability_corr_cov.html", ""),
    (SEC_PROB, "19 Distributions", BASE + "19_probability_distributions.html", ""),
    (SEC_PROB, "20 Convergence, LLN, CLT", BASE + "20_probability_convergence_modes_lln_clt.html", ""),

    (SEC_STAT, "21 Statistics fundamentals", BASE + "21_stat_fundamentals.html", "Essential."),
    (SEC_STAT, "22 Estimators", BASE + "22_stat_estimators.html",
     "First part is important. Fisher information and Cramer-Rao are not super important."),
    (SEC_STAT, "23 MLE & MAP", BASE + "23_stat_mle_map.html", "Both MLE and MAP are important."),
    (SEC_STAT, "24 Confidence intervals", BASE + "24_stat_confidence_intervals.html",
     "The sampling-related ideas and intuition matter; the calculations don't."),
    (SEC_STAT, "25 Hypothesis testing", BASE + "25_stat_hypothesis_testing.html",
     "Good to understand the intuition; the rest isn't super important."),
    (SEC_STAT, "26 Classical tests", BASE + "26_stat_classical_tests.html", "Very optional."),
    (SEC_STAT, "27 How to lie with statistics", BASE + "27_stat_how_to_lie.html", "Just for fun."),

    (SEC_INFO, "28 Information theory (entropy, KL)", BASE + "28_info_theory_entropy_kl.html",
     "More of a bonus, but would be nice to watch."),
    (SEC_INFO, "30 Curse of dimensionality", BASE + "30_curse_of_dimensionality.html",
     "Video not available yet. Good for building the intuition you'll need later."),
]


def main() -> None:
    wb = load_workbook(XLSX)

    # drop scratch + any prior version of the overview sheet (idempotency).
    # Match any "Probability, Stats..." variant, incl. Google-Sheets-escaped "&amp;".
    for name in list(wb.sheetnames):
        if name == "Sheet2" or name.startswith("Probability, Stats"):
            del wb[name]

    ws = wb.create_sheet(SHEET_NAME)

    # header (bold, matching the current state of the sibling sheets)
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # data rows
    for i, (section, item, link, note) in enumerate(ROWS, start=2):
        fill = PatternFill("solid", fgColor=SECTION_COLOR[section])

        ws.cell(row=i, column=1, value=section)
        ws.cell(row=i, column=2, value=item)

        link_cell = ws.cell(row=i, column=3, value=link)
        if link:
            link_cell.hyperlink = link
            link_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=i, column=4, value=note)
        ws.cell(row=i, column=5, value="")  # Done (add checkboxes in Sheets)
        ws.cell(row=i, column=6, value="")  # Question / comment

        for col in range(1, 7):
            c = ws.cell(row=i, column=col)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # column widths (close to the sibling sheets)
    widths = {1: 36, 2: 34, 3: 48, 4: 60, 5: 15, 6: 45}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    wb.save(XLSX)
    log.info(f"Added '{SHEET_NAME}' ({len(ROWS)} rows) to {XLSX.resolve()}")
    log.info(f"Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
